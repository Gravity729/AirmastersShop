import openpyxl
import os

# Путь к вашему Excel-файлу
file_path = r"C:\Users\Фёдор\Desktop\Orders.xlsx"

# Создаём или открываем файл
if not os.path.exists(file_path):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    # Добавляем заголовки
    sheet.append(["Номер заказа", "Имя клиента", "Номер телефона", "Почта клиента",
                  "Страна", "Город", "Названия товаров", "Количество", "Итоговая стоимость"])
else:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

# Тестовая запись
order_number = sheet.max_row  # Номер заказа = номер строки
sheet.append([
    order_number, "Тестовый клиент", "123456789", "test@mail.com",
    "Россия", "Москва", "Товар 1, Товар 2", "2, 3", 5000
])

# Сохраняем файл
workbook.save(file_path)

print(f"Файл успешно обновлен! Путь: {file_path}")
