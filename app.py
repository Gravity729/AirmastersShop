import re
from flask import Flask, render_template, request, jsonify, redirect, session, url_for
import sqlite3
import bcrypt
import os
import base64
from io import BytesIO
from PIL import Image, ExifTags
from openpyxl import load_workbook, Workbook


app = Flask(__name__)

# Генерация случайного секретного ключа для сессий
app.secret_key = os.urandom(24)

# Путь к Excel-файлу
EXCEL_FILE_PATH = r"C:\Users\Фёдор\Desktop\Orders.xlsx"

@app.route('/order', methods=['GET', 'POST'])
def order():
    if 'user_id' not in session:
        return redirect(url_for('login'))  # Перенаправить на логин, если пользователь не авторизован

    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor()

    # Получаем данные пользователя
    cursor.execute("SELECT fName, phone, email FROM User WHERE userId = ?", (user_id,))
    user = cursor.fetchone()

    # Получаем товары из корзины
    cursor.execute("""
        SELECT Cart.productId, Product.name, Product.price, Product.stockQuantity, Cart.quantity,
               (SELECT photoUrl FROM ProductPhoto WHERE ProductPhoto.productId = Product.productId LIMIT 1) AS photoUrl
        FROM Cart
        JOIN Product ON Cart.productId = Product.productId
        WHERE Cart.userId = ?
    """, (user_id,))

    cart_items = []
    total_cost = 0

    # Проверяем количество и фильтруем товары
    for item in cursor.fetchall():
        if item['stockQuantity'] > 0:
            adjusted_quantity = min(item['quantity'], item['stockQuantity'])
            total_cost += adjusted_quantity * item['price']
            cart_items.append({
                'productId': item['productId'],
                'name': item['name'],
                'price': item['price'],
                'quantity': adjusted_quantity,
                'photoUrl': item['photoUrl'],
            })

    conn.close()

    if request.method == 'POST':
        name = request.form['name']
        phone = request.form['phone']
        email = request.form['email']
        country = request.form['country']
        city = request.form['city']

        # Работа с Excel-файлом
        try:
            # Если файла нет – создаём его с заголовками
            if not os.path.exists(EXCEL_FILE_PATH):
                wb = Workbook()
                sheet = wb.active
                sheet.append(["№ Заказа", "Имя", "Телефон", "Email", "Страна", "Город", "Товары", "Кол-во", "Сумма"])
                wb.save(EXCEL_FILE_PATH)
            else:
                wb = load_workbook(EXCEL_FILE_PATH)
                sheet = wb.active

            # Номер заказа и добавление данных
            order_number = sheet.max_row  # Номер заказа
            products = ",".join([item['name'] for item in cart_items])
            quantities = ",".join([str(item['quantity']) for item in cart_items])

            sheet.append([order_number, name, phone, email, country, city, products, quantities, total_cost])
            wb.save(EXCEL_FILE_PATH)

        except PermissionError:
            return """
            <script>
                alert("Ошибка доступа к файлу! Закройте файл и попробуйте снова.");
                window.history.back();
            </script>
            """
        except Exception as e:
            return f"Произошла ошибка при работе с файлом: {e}"

        # Обновляем количество товаров на складе
        conn = get_db_connection()
        cursor = conn.cursor()
        for item in cart_items:
            cursor.execute("""
                UPDATE Product
                SET stockQuantity = stockQuantity - ?
                WHERE productId = ?
            """, (item['quantity'], item['productId']))
        conn.commit()
        conn.close()

        session.pop('cart', None)

        # Возвращаем JavaScript для всплывающего окна
        return f"""
        <script>
            alert("Ваш заказ успешно оформлен! Номер вашего заказа: №{order_number}");
            window.location.href = "/";
        </script>
        """

    # Передаём данные пользователя и корзины
    return render_template('order.html', order_items=cart_items, total_cost=total_cost, user=user)


# Функция подключения к базе данных
def get_db_connection():
    conn = sqlite3.connect('airmasters.db')
    conn.row_factory = sqlite3.Row
    return conn




ALLOWED_EXTENSIONS = {'jpg', 'jpeg', 'png'}

# Проверка допустимых расширений файлов
def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# Функция для коррекции ориентации изображения на основе EXIF
def correct_image_orientation(img):
    try:
        # Получаем EXIF данные
        exif = img._getexif()
        if exif:
            # Найдём тег для ориентации
            orientation_key = next(
                (tag for tag, value in ExifTags.TAGS.items() if value == "Orientation"), None
            )

            if orientation_key:
                orientation = exif.get(orientation_key)
                # Проверяем ориентацию и корректируем изображение
                if orientation == 3:  # Перевёрнуто на 180
                    img = img.rotate(180, expand=True)
                elif orientation == 6:  # Повернуто на 270° по часовой стрелке
                    img = img.rotate(270, expand=True)
                elif orientation == 8:  # Повернуто на 90° против часовой стрелки
                    img = img.rotate(90, expand=True)
    except Exception as e:
        # Если что-то пошло не так, просто возвращаем изображение как есть
        print(f"Ошибка обработки EXIF: {e}")
    return img


@app.route('/getModels/<int:brand_id>', methods=['GET'])
def get_models(brand_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT carModelId, name FROM CarModel WHERE carBrandId = ?", (brand_id,))
        models = cursor.fetchall()
        return jsonify([{"carModelId": row["carModelId"], "name": row["name"]} for row in models])
    except sqlite3.Error as e:
        print(f"Ошибка базы данных: {e}")
        return jsonify([])
    finally:
        conn.close()


@app.route('/addProduct', methods=['GET', 'POST'])
def add_product():
    if 'user_id' not in session or session.get('user_role') != 1:
        return redirect(url_for('login'))

    errors = {}
    conn = get_db_connection()
    cursor = conn.cursor()

    # Получаем список марок и моделей для выпадающих списков
    cursor.execute("SELECT carBrandId, name FROM CarBrand")
    car_brands = cursor.fetchall()

    cursor.execute("SELECT carModelId, name, carBrandId FROM CarModel")
    car_models = cursor.fetchall()

    if request.method == 'POST':
        # Получение данных из формы
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        price = request.form.get('price', '').strip()
        stock_quantity = request.form.get('stockQuantity', '').strip()
        car_brand_id = request.form.get('brand', '').strip()
        car_model_id = request.form.get('model', '').strip()
        category_id = request.form.get('categoryId', '')
        photos = request.files.getlist('photos')

        # Проверка полей
        if not name:
            errors['name'] = "Название товара обязательно для заполнения."
        if not description:
            errors['description'] = "Описание обязательно для заполнения."
        try:
            price = int(price)
            if price <= 0:
                errors['price'] = "Цена должна быть строго больше 0."
        except ValueError:
            errors['price'] = "Цена должна быть целым числом."
        try:
            stock_quantity = int(stock_quantity)
            if stock_quantity < 0:
                errors['stockQuantity'] = "Количество не может быть отрицательным."
        except ValueError:
            errors['stockQuantity'] = "Количество должно быть целым числом."

        # Проверка фотографий
        if not photos or all(photo.filename.strip() == '' for photo in photos):
            errors['photos'] = "Товар должен содержать хотя бы одну фотографию."

        # Если есть ошибки
        if errors:
            return render_template(
                'addProduct.html',
                errors=errors,
                name=name,
                description=description,
                price=price,
                stockQuantity=stock_quantity,
                category_id=category_id,  # Передача обратно текущей категории
                car_brands=car_brands,
                car_models=car_models,
                selected_brand=str(car_brand_id),
                selected_model=str(car_model_id)
            )

        # Сохранение данных в Product
        cursor.execute("""
            INSERT INTO Product (name, description, price, stockQuantity, carModelId, categoryId)
            VALUES (?, ?, ?, ?, ?, ?)
        """, (name, description, price, stock_quantity, car_model_id or None, category_id))
        product_id = cursor.lastrowid

        # Сохранение фотографий
        for photo in photos:
            if photo and allowed_file(photo.filename):
                img = Image.open(photo)
                img = correct_image_orientation(img)
                img = img.convert("RGB")

                buffer = BytesIO()
                img.save(buffer, format="JPEG", quality=95)
                img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
                photo_url = f"data:image/jpeg;base64,{img_base64}"

                cursor.execute("""
                    INSERT INTO ProductPhoto (productId, photoUrl)
                    VALUES (?, ?)
                """, (product_id, photo_url))

        conn.commit()
        conn.close()
        return redirect(url_for('products', category=category_id))

    # GET-запрос
    conn.close()
    category_id = request.args.get('category', '')
    # GET-запрос (без ошибок)
    return render_template(
        'addProduct.html',
        errors={},
        name='',
        description='',
        price='',
        stockQuantity='',
        category_id=category_id,
        car_brands=car_brands,
        car_models=car_models,
        selected_brand='',  # Пустое значение по умолчанию
        selected_model=''  # Пустое значение по умолчанию
    )


@app.route('/editProduct/<int:product_id>', methods=['GET', 'POST'])
def edit_product(product_id):
    if 'user_id' not in session or session.get('user_role') != 1:
        return redirect(url_for('login'))

    errors = {}
    conn = get_db_connection()
    cursor = conn.cursor()

    # Получаем текущие данные о товаре
    cursor.execute("""
        SELECT Product.*, CarModel.carBrandId 
        FROM Product
        LEFT JOIN CarModel ON Product.carModelId = CarModel.carModelId
        WHERE Product.productId = ?
    """, (product_id,))
    product = cursor.fetchone()

    # Получаем текущие фотографии
    cursor.execute("SELECT photoId, photoUrl FROM ProductPhoto WHERE productId = ?", (product_id,))
    photos = cursor.fetchall()

    # Получаем марки и модели
    cursor.execute("SELECT carBrandId, name FROM CarBrand")
    car_brands = cursor.fetchall()
    cursor.execute("SELECT carModelId, name, carBrandId FROM CarModel")
    car_models = cursor.fetchall()

    if request.method == 'POST':
        # Получаем данные формы
        name = request.form.get('name', '').strip()
        description = request.form.get('description', '').strip()
        price = request.form.get('price', '').strip()
        stock_quantity = request.form.get('stockQuantity', '').strip()
        car_brand_id = request.form.get('brand', '').strip()
        car_model_id = request.form.get('model', '').strip()
        new_photos = request.files.getlist('photos')
        delete_photo_ids = request.form.getlist('delete_photos')  # Чекбоксы удаления

        # Валидация
        if not name:
            errors['name'] = "Поле 'Название' обязательно для заполнения."
        if not description:
            errors['description'] = "Поле 'Описание' обязательно для заполнения."
        if not price or not price.isdigit() or int(price) <= 0:
            errors['price'] = "Поле 'Стоимость' должно быть числом больше 0."
        if not stock_quantity or not stock_quantity.isdigit() or int(stock_quantity) < 0:
            errors['stockQuantity'] = "Поле 'Количество' должно быть положительным числом."

        # Удаляем фотографии, если есть отмеченные чекбоксы
        if delete_photo_ids:
            remaining_photos = len(photos) - len(delete_photo_ids)
            if remaining_photos <= 0 and not new_photos:
                errors['photos'] = "Нельзя удалить все фотографии. Добавьте хотя бы одну новую фотографию."
            else:
                for photo_id in delete_photo_ids:
                    cursor.execute("DELETE FROM ProductPhoto WHERE photoId = ?", (photo_id,))

        # Сохраняем новые фотографии
        for photo in new_photos:
            if photo and allowed_file(photo.filename):
                try:
                    img = Image.open(photo)
                    img = correct_image_orientation(img)
                    img = img.convert("RGB")  # Преобразование в RGB

                    # Сохраняем изображение в base64
                    buffer = BytesIO()
                    img.save(buffer, format="JPEG", quality=95)
                    img_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
                    photo_url = f"data:image/jpeg;base64,{img_base64}"

                    # Добавляем ссылку в базу данных
                    cursor.execute("""
                        INSERT INTO ProductPhoto (productId, photoUrl)
                        VALUES (?, ?)
                    """, (product_id, photo_url))
                except Exception as e:
                    print(f"Ошибка при обработке фотографии: {e}")

        # Обновляем товар
        if not errors:
            cursor.execute("""
                UPDATE Product
                SET name = ?, description = ?, price = ?, stockQuantity = ?, carModelId = ?
                WHERE productId = ?
            """, (name, description, price, stock_quantity, car_model_id or None, product_id))
            conn.commit()
            conn.close()
            return redirect(url_for('product_info', product_id=product_id))

    conn.close()
    return render_template(
        'editProduct.html',
        product=product,
        photos=photos,
        car_brands=car_brands,
        car_models=car_models,
        errors=errors
    )




# Подключение к базе данных
def get_db_connection():
    conn = sqlite3.connect('airmasters.db')
    conn.row_factory = sqlite3.Row  # Позволяет обращаться к колонкам через имена
    return conn


# Проверка входа пользователя
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        login_input = request.form['login']
        password = request.form['password']

        # Поиск пользователя в базе данных
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT * FROM User WHERE login = ?", (login_input,))
        user = cursor.fetchone()
        conn.close()

        # Проверка пароля
        if user and bcrypt.checkpw(password.encode('utf-8'), user['password']):
            # Установка данных в сессию
            session['user_id'] = user['userId']
            session['user_name'] = user['fName']
            session['user_role'] = user['role']
            session['user_phone'] = user['phone']
            session['user_email'] = user['email']

            return jsonify({
                "status": "success",
                "message": f"Добро пожаловать, {user['fName']}!",
                "redirect_url": url_for('lk')  # URL для перенаправления на личный кабинет
            })
        else:
            return jsonify({
                "status": "error",
                "message": "Неверный логин или пароль"
            })

    return render_template('authorization.html')


# Личный кабинет
@app.route('/lk')
def lk():
    if 'user_id' not in session:
        return redirect(url_for('authorization'))

    return render_template(
        'lk.html',
        user_name=session.get('user_name', 'Пользователь'),
        user_role=session.get('user_role', 2),
        user_phone=session.get('user_phone', ''),
        user_email=session.get('user_email', '')
    )


# Выход из аккаунта
@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('authorization'))


# Регистрация нового пользователя
@app.route('/register', methods=['GET', 'POST'])
def register():
    if request.method == 'POST':
        # Получение данных из формы
        name = request.form.get('name', '').strip()
        phone = request.form.get('phone', '').strip()
        email = request.form.get('email', '').strip()
        login = request.form.get('login', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()

        # Словарь для ошибок
        errors = {}

        # Проверка полей
        if not name:
            errors['name'] = "Имя обязательно для заполнения."
        if not phone.startswith('+') or not phone[1:].isdigit() or len(phone) < 10:
            errors['phone'] = "Телефон должен быть в формате +79990001122."
        if not email:
            errors['email'] = "E-mail обязателен для заполнения."
        if not login:
            errors['login'] = "Логин обязателен для заполнения."

        # Регулярное выражение для проверки пароля
        password_regex = re.compile(r'^(?=.*[A-Z])(?=.*\d)(?=.*[-_!#*]).{5,}$')

        if not password_regex.match(password):
            errors['password'] = (
                "Пароль должен содержать минимум 5 символов, "
                "одну заглавную букву, одну цифру и один из символов -_!#*."
            )
        if password != confirm_password:
            errors['confirm_password'] = "Пароли не совпадают."

        # Проверка уникальности логина и email
        if not errors:
            try:
                conn = get_db_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT * FROM User WHERE login = ? OR email = ?", (login, email))
                existing_user = cursor.fetchone()
                if existing_user:
                    if existing_user['login'] == login:
                        errors['login'] = "Пользователь с таким логином уже существует."
                    if existing_user['email'] == email:
                        errors['email'] = "Пользователь с таким email уже существует."
                else:
                    hashed_password = bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt())
                    cursor.execute("""
                        INSERT INTO User (fName, phone, email, login, password, role)
                        VALUES (?, ?, ?, ?, ?, 2)
                    """, (name, phone, email, login, hashed_password))
                    conn.commit()
                    conn.close()
                    return render_template('registration.html', success="Вы успешно зарегистрировались!", errors={})

            except sqlite3.Error as e:
                errors['database'] = f"Ошибка базы данных: {e}"

        return render_template(
            'registration.html',
            errors=errors,
            name=name,
            phone=phone,
            email=email,
            login=login
        )

    return render_template('registration.html', errors={}, name='', phone='', email='', login='')


# Получение списка категорий
def get_categories():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        try:
            cursor.execute("SELECT categoryId, name FROM Category")
            categories = cursor.fetchall()
            return [{"id": category["categoryId"], "name": category["name"]} for category in categories]
        except sqlite3.Error as e:
            print(f"Ошибка базы данных: {e}")
            return []


# Получение списка марок автомобилей
def get_car_brands():
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT carBrandId, name FROM CarBrand")
        brands = cursor.fetchall()
        return [{"carBrandId": brand["carBrandId"], "name": brand["name"]} for brand in brands]
    except sqlite3.Error as e:
        print(f"Ошибка базы данных: {e}")
        return []
    finally:
        conn.close()


# Получение списка моделей автомобилей по марке
def get_car_models(brand_id):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        print(f"Получение моделей для бренда с ID: {brand_id}")  # Отладка
        cursor.execute("SELECT carModelId, name FROM CarModel WHERE carBrandId = ?", (brand_id,))
        models = cursor.fetchall()
        print(f"Найденные модели: {models}")  # Отладка
        return [{"modelId": row["carModelId"], "name": row["name"]} for row in models]
    except sqlite3.Error as e:
        print(f"Ошибка базы данных: {e}")
        return []
    finally:
        conn.close()


# Получение списка товаров по категории
def get_products_by_category(category_id=None):
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        if category_id:
            query = """
                SELECT DISTINCT Product.productId, Product.name, Product.description, Product.price, 
                       (SELECT photoUrl FROM ProductPhoto WHERE ProductPhoto.productId = Product.productId LIMIT 1) AS photoUrl
                FROM Product
                WHERE Product.categoryId = ?
            """
            cursor.execute(query, (category_id,))
        else:
            query = """
                SELECT DISTINCT Product.productId, Product.name, Product.description, Product.price, 
                       (SELECT photoUrl FROM ProductPhoto WHERE ProductPhoto.productId = Product.productId LIMIT 1) AS photoUrl
                FROM Product
            """
            cursor.execute(query)

        return [
            {"id": row["productId"], "name": row["name"], "description": row["description"], "price": row["price"], "photoUrl": row["photoUrl"]}
            for row in cursor.fetchall()
        ]
    except sqlite3.Error as e:
        print(f"Ошибка базы данных: {e}")
        return []
    finally:
        conn.close()


# Внедрение списка категорий в шаблоны
@app.context_processor
def inject_categories():
    return {"categories": get_categories()}


# Главная страница
@app.route('/')
def home():
    return render_template('home.html')


# Страница "Корзина"
@app.route('/cart')
def cart():
    if 'user_id' not in session:  # Проверка авторизации
        return render_template('cart.html', cart_items=[], total_items=0, total_cost=0)

    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
        SELECT Product.productId, Product.name, Product.price, Product.stockQuantity,
               (SELECT photoUrl FROM ProductPhoto WHERE ProductPhoto.productId = Product.productId LIMIT 1) AS photoUrl,
               IFNULL(Cart.quantity, 0) AS cartQuantity
        FROM Cart
        JOIN Product ON Cart.productId = Product.productId
        WHERE Cart.userId = ?
    """
    cursor.execute(query, (user_id,))
    cart_items = [dict(row) for row in cursor.fetchall()]  # Преобразование в словарь

    # Получаем товары, которые в избранном
    cursor.execute("SELECT productId FROM Favorite WHERE userId = ?", (user_id,))
    favorite_products = {row['productId'] for row in cursor.fetchall()}

    # Добавляем флаг is_favorite для каждого товара
    for item in cart_items:
        item['is_favorite'] = 1 if item['productId'] in favorite_products else 0

    conn.close()

    # Подсчет итогов
    total_items = sum(item['cartQuantity'] for item in cart_items if item['stockQuantity'] > 0)
    total_cost = sum(item['price'] * item['cartQuantity'] for item in cart_items if item['stockQuantity'] > 0)

    return render_template(
        'cart.html',
        cart_items=cart_items,
        total_items=total_items,
        total_cost=total_cost
    )


@app.route('/favorite')
def favorites():
    if 'user_id' not in session:  # Проверка авторизации
        return redirect(url_for('authorization'))

    user_id = session['user_id']
    conn = get_db_connection()
    cursor = conn.cursor()

    # Получаем избранные товары с данными о наличии на складе и первой фотографией
    query = """
        SELECT Product.productId, Product.name, Product.price, Product.stockQuantity, 
               (SELECT photoUrl FROM ProductPhoto WHERE ProductPhoto.productId = Product.productId LIMIT 1) AS photoUrl, 
               IFNULL(Cart.quantity, 0) AS cartQuantity,
               1 AS is_favorite  -- Указываем, что товар уже в избранном
        FROM Favorite
        JOIN Product ON Favorite.productId = Product.productId
        LEFT JOIN Cart ON Cart.productId = Product.productId AND Cart.userId = ?
        WHERE Favorite.userId = ?
    """
    cursor.execute(query, (user_id, user_id))
    favorite_items = cursor.fetchall()
    conn.close()

    return render_template(
        'favorite.html',
        favorite_items=favorite_items
    )


@app.route('/add_to_favorite', methods=['POST'])
def add_to_favorite():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Пользователь не авторизован'})

    user_id = session['user_id']
    data = request.get_json()
    product_id = data.get('product_id')

    if not product_id:
        return jsonify({'status': 'error', 'message': 'Некорректный ID товара'})

    conn = get_db_connection()
    cursor = conn.cursor()

    # Проверяем, есть ли запись уже в избранном
    cursor.execute("SELECT * FROM Favorite WHERE userId = ? AND productId = ?", (user_id, product_id))
    if not cursor.fetchone():
        cursor.execute("INSERT INTO Favorite (userId, productId) VALUES (?, ?)", (user_id, product_id))
        conn.commit()

    conn.close()
    return jsonify({'status': 'success'})


@app.route('/remove_from_favorite', methods=['POST'])
def remove_from_favorite():
    if 'user_id' not in session:
        return jsonify({'status': 'error', 'message': 'Пользователь не авторизован'})

    user_id = session['user_id']
    data = request.get_json()
    product_id = data.get('product_id')

    if not product_id:
        return jsonify({'status': 'error', 'message': 'Некорректный ID товара'})

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("DELETE FROM Favorite WHERE userId = ? AND productId = ?", (user_id, product_id))
    conn.commit()
    conn.close()

    return jsonify({'status': 'success'})


@app.route('/update_quantity', methods=['POST'])
def update_quantity():
    if 'user_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    product_id = request.args.get('product_id', type=int)
    delta = request.args.get('delta', type=int)
    user_id = session['user_id']

    conn = get_db_connection()
    cursor = conn.cursor()

    # Проверка доступного количества товара
    cursor.execute("SELECT stockQuantity FROM Product WHERE productId = ?", (product_id,))
    product = cursor.fetchone()
    if not product:
        return jsonify({"error": "Product not found"}), 404

    stock_quantity = product['stockQuantity']

    # Получение текущего количества товара в корзине
    cursor.execute("SELECT quantity FROM Cart WHERE userId = ? AND productId = ?", (user_id, product_id))
    cart_item = cursor.fetchone()
    if not cart_item:
        return jsonify({"error": "Cart item not found"}), 404

    current_quantity = cart_item['quantity']
    new_quantity = current_quantity + delta

    # Ограничение: нельзя добавить больше, чем есть на складе
    if new_quantity > stock_quantity:
        return jsonify({
            "error": f"Cannot add more than {stock_quantity} items. Available stock: {stock_quantity}."
        }), 400

    # Ограничение: количество не может быть отрицательным
    if new_quantity < 0:
        return jsonify({"error": "Quantity cannot be negative"}), 400

    # Удаление товара, если количество становится 0
    if new_quantity == 0:
        cursor.execute("DELETE FROM Cart WHERE userId = ? AND productId = ?", (user_id, product_id))
    else:
        # Обновление количества в корзине
        cursor.execute("""
            UPDATE Cart
            SET quantity = ?
            WHERE userId = ? AND productId = ?
        """, (new_quantity, user_id, product_id))
    conn.commit()
    conn.close()

    return jsonify({"status": "success", "new_quantity": new_quantity})


@app.route('/remove_item', methods=['POST'])
def remove_item():
    if 'user_id' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    product_id = request.args.get('product_id', type=int)
    user_id = session['user_id']

    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute("""
        DELETE FROM Cart
        WHERE userId = ? AND productId = ?
    """, (user_id, product_id))
    conn.commit()
    conn.close()

    return jsonify({"status": "success"})


# Страница "О нас"
@app.route('/about')
def about():
    return render_template('about.html')


# Страница "Контакты"
@app.route('/contacts')
def contacts():
    return render_template('contacts.html')


# Страница товаров
@app.route('/products')
def products():
    category_id = request.args.get('category', type=int)
    selected_brand_id = request.args.get('brand', type=int)
    selected_model_id = request.args.get('model', type=int)
    user_id = session.get('user_id', None)

    conn = get_db_connection()
    cursor = conn.cursor()

    # Базовый SQL-запрос с проверкой избранного
    query = """
        SELECT Product.productId, Product.name, Product.description, Product.price, 
       Product.stockQuantity,  -- Добавлен столбец
       Product.carModelId, CarModel.carBrandId, 
       (SELECT photoUrl FROM ProductPhoto WHERE ProductPhoto.productId = Product.productId LIMIT 1) AS photoUrl,
       CASE WHEN EXISTS (
           SELECT 1 FROM Favorite 
           WHERE Favorite.productId = Product.productId AND Favorite.userId = ?
       ) THEN 1 ELSE 0 END AS is_favorite
       FROM Product
        LEFT JOIN CarModel ON Product.carModelId = CarModel.carModelId
    """
    params = [user_id]  # Передаем user_id

    # Добавляем условия фильтрации
    conditions = []
    if category_id:
        conditions.append("Product.categoryId = ?")
        params.append(category_id)
    if selected_brand_id:
        conditions.append("CarModel.carBrandId = ?")
        params.append(selected_brand_id)
    if selected_model_id:
        conditions.append("Product.carModelId = ?")
        params.append(selected_model_id)

    if conditions:
        query += " WHERE " + " AND ".join(conditions)

    # Выполняем запрос
    cursor.execute(query, tuple(params))
    products = [
        {
            "id": row["productId"],
            "name": row["name"],
            "description": row["description"],
            "price": row["price"],
            "stockQuantity": row["stockQuantity"],  # Добавьте эту строку
            "photoUrl": row["photoUrl"] or "placeholder.jpg",
            "carBrandId": row["carBrandId"],
            "carModelId": row["carModelId"],
            "is_favorite": row["is_favorite"] == 1  # Преобразуем в True/False
        }
        for row in cursor.fetchall()
    ]

    car_brands = get_car_brands()
    cart_items = {}
    if user_id:
        cursor.execute("""
            SELECT productId, quantity
            FROM Cart
            WHERE userId = ?
        """, (user_id,))
        cart_items = {item['productId']: item['quantity'] for item in cursor.fetchall()}

    conn.close()

    return render_template(
        'products.html',
        products=products,
        selected_category=category_id,
        car_brands=car_brands,
        cart_items=cart_items
    )


# Подробная информация о товаре
@app.route('/product/<int:product_id>')
def product_info(product_id):
    conn = get_db_connection()
    cursor = conn.cursor()

    # Получение информации о товаре
    cursor.execute("SELECT * FROM Product WHERE productId = ?", (product_id,))
    product = cursor.fetchone()

    # Проверка, существует ли товар в базе
    if not product:
        conn.close()
        return "Товар не найден", 404

    # Получение категории товара
    cursor.execute("SELECT categoryId FROM Product WHERE productId = ?", (product_id,))
    category = cursor.fetchone()

    # Получение фотографий товара
    cursor.execute("SELECT photoUrl FROM ProductPhoto WHERE productId = ?", (product_id,))
    photos = [row['photoUrl'] for row in cursor.fetchall()]

    cart_items = {}
    is_favorite = False  # По умолчанию товар не в избранном

    # Проверяем авторизацию пользователя
    if 'user_id' in session:
        user_id = session['user_id']
        cursor.execute("SELECT productId, quantity FROM Cart WHERE userId = ?", (user_id,))
        cart_items = {item['productId']: item['quantity'] for item in cursor.fetchall()}

        cursor.execute("SELECT 1 FROM Favorite WHERE userId = ? AND productId = ?", (user_id, product_id))
        is_favorite = cursor.fetchone() is not None

    conn.close()

    # Возвращаем подробности о товаре
    product_data = {
        'productId': product['productId'],
        'name': product['name'],
        'price': product['price'],
        'description': product['description'],
        'stockQuantity': product['stockQuantity'],
        'photos': photos
    }
    return render_template(
        'infoProduct.html',
        product=product_data,
        cart_items=cart_items,
        selected_category=category['categoryId'] if category else None,
        is_favorite=is_favorite
    )


# Страница авторизации
@app.route('/authorization')
def authorization():
    return render_template('authorization.html')


# Страница регистрации
@app.route('/registration')
def registration():
    return render_template('registration.html')


# Страница регистрации
@app.route('/addProduct')
def addProduct():
    return render_template('addProduct.html')


# API для получения списка моделей
@app.route('/api/models', methods=['GET'])
def api_models():
    brand_id = request.args.get('brandId', type=int)
    if not brand_id:
        return jsonify({"error": "Missing brandId"}), 400

    models = get_car_models(brand_id)
    return jsonify(models)


# API для получения списка товаров
@app.route('/api/products')
def api_products():
    category_id = request.args.get('category', type=int)
    products = get_products_by_category(category_id)
    return jsonify(products)


# Добавление товара в корзину
@app.route('/add_to_cart', methods=['POST'])
def add_to_cart():
    if 'user_id' not in session:
        return jsonify({
            "status": "error",
            "message": "Пожалуйста, авторизуйтесь, чтобы добавить товары в корзину."
        }), 401

    product_id = request.json.get('product_id')
    quantity = request.json.get('quantity', 1)
    user_id = session['user_id']

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT stockQuantity FROM Product WHERE productId = ?", (product_id,))
    product = cursor.fetchone()
    if not product:
        conn.close()
        return jsonify({"status": "error", "message": "Товар не найден."}), 404

    if product['stockQuantity'] < quantity:
        conn.close()
        return jsonify({
            "status": "error",
            "message": f"Недостаточно товара на складе. Доступно: {product['stockQuantity']}."
        }), 400

    cursor.execute("SELECT quantity FROM Cart WHERE userId = ? AND productId = ?", (user_id, product_id))
    cart_item = cursor.fetchone()

    if cart_item:
        new_quantity = cart_item['quantity'] + quantity
        if new_quantity > product['stockQuantity']:
            conn.close()
            return jsonify({
                "status": "error",
                "message": f"Невозможно добавить более {product['stockQuantity']} единиц товара."
            }), 400
        cursor.execute("""
            UPDATE Cart
            SET quantity = ?
            WHERE userId = ? AND productId = ?
        """, (new_quantity, user_id, product_id))
    else:
        cursor.execute("""
            INSERT INTO Cart (userId, productId, quantity)
            VALUES (?, ?, ?)
        """, (user_id, product_id, quantity))

    conn.commit()
    cursor.execute("SELECT quantity FROM Cart WHERE userId = ? AND productId = ?", (user_id, product_id))
    updated_quantity = cursor.fetchone()['quantity']
    conn.close()

    return jsonify({"status": "success", "quantity": updated_quantity})


@app.route('/remove_from_cart', methods=['POST'])
def remove_from_cart():
    if 'user_id' not in session:
        return jsonify({
            "status": "error",
            "message": "Пожалуйста, авторизуйтесь, чтобы удалить товары из корзины."
        }), 401

    product_id = request.json.get('product_id')
    user_id = session['user_id']

    conn = get_db_connection()
    cursor = conn.cursor()

    cursor.execute("SELECT quantity FROM Cart WHERE userId = ? AND productId = ?", (user_id, product_id))
    cart_item = cursor.fetchone()

    if not cart_item:
        conn.close()
        return jsonify({"status": "error", "message": "Товар не найден в корзине."}), 404

    new_quantity = cart_item['quantity'] - 1

    if new_quantity == 0:
        cursor.execute("DELETE FROM Cart WHERE userId = ? AND productId = ?", (user_id, product_id))
    else:
        cursor.execute("""
            UPDATE Cart
            SET quantity = ?
            WHERE userId = ? AND productId = ?
        """, (new_quantity, user_id, product_id))

    conn.commit()
    conn.close()

    return jsonify({"status": "success", "quantity": new_quantity})


if __name__ == "__main__":
    app.run(debug=True, port=5001)
